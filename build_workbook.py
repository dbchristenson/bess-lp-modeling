"""
Generate Excel workbook from BESS + DR optimization results.

Reads results.pkl exported by build_model.py and produces Excel workbook.

Run: uv run python build_workbook.py
"""

import pickle
from datetime import datetime, timedelta
from pathlib import Path

import numpy as np
import openpyxl
from openpyxl.styles import Alignment, Border, Font, NamedStyle, PatternFill, Side
from openpyxl.utils import get_column_letter

SCRIPT_DIR = Path(__file__).resolve().parent
OUTPUT_XLSX = SCRIPT_DIR / "BESS_DR_Model_Results.xlsx"
RESULTS_PKL = SCRIPT_DIR / "results.pkl"

T = 8760
YEAR = 2023
D_MW = 10.0
CARBON_INTENSITY = 225

BESS_UPDATED = {
    "label": "NREL ATB 2024 Adv.",
    "c_P_cap": 221.4,
    "c_E_cap": 239.6,
    "c_P_opex": 5.5,
    "c_E_opex": 6.0,
}

WACC = 0.08
LT_BESS = 15
AF_BESS = WACC * (1 + WACC) ** LT_BESS / ((1 + WACC) ** LT_BESS - 1)

P_OPTIONS = [2, 4, 6, 8, 10]
TAU_OPTIONS = [1, 2, 4]

MONTH_NAMES = ["January", "February", "March", "April", "May", "June",
               "July", "August", "September", "October", "November", "December"]

# ---------------------------------------------------------------------------
# Styles
# ---------------------------------------------------------------------------

HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
SUBHEADER_FILL = PatternFill(start_color="D6DCE4", end_color="D6DCE4", fill_type="solid")
TOTAL_FILL = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
HIGHLIGHT_FILL = PatternFill(start_color="D4EDDA", end_color="D4EDDA", fill_type="solid")
BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)
BORDER_TOP_THICK = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="medium"),
    bottom=Side(style="thin"),
)

NUM_FMT_INT = '#,##0'
NUM_FMT_DEC1 = '#,##0.0'
NUM_FMT_DEC2 = '#,##0.00'
NUM_FMT_PCT = '0.0%'
NUM_FMT_YR = '0.0'

TAB_BLUE = "2F5496"
TAB_GREEN = "548235"
TAB_ORANGE = "C65911"
TAB_PURPLE = "7030A0"


def hour_index_to_datetime(h):
    return datetime(YEAR, 1, 1) + timedelta(hours=int(h))


def build_calendar():
    months = np.empty(T, dtype=int)
    hours = np.empty(T, dtype=int)
    dows = np.empty(T, dtype=int)
    days = np.empty(T, dtype=int)
    for h in range(T):
        dt = hour_index_to_datetime(h)
        months[h] = dt.month
        hours[h] = dt.hour
        dows[h] = dt.weekday()
        days[h] = h // 24
    return months, hours, dows, days


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def style_header_row(ws, row, ncols):
    for col in range(1, ncols + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = BORDER


def style_data_cell(ws, row, col, fmt=None):
    cell = ws.cell(row=row, column=col)
    cell.border = BORDER
    cell.alignment = Alignment(horizontal="center")
    if fmt:
        cell.number_format = fmt


def style_total_row(ws, row, ncols, fmt_map=None):
    for col in range(1, ncols + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = TOTAL_FILL
        cell.font = Font(bold=True)
        cell.border = BORDER_TOP_THICK
        cell.alignment = Alignment(horizontal="center")
        if fmt_map and col in fmt_map:
            cell.number_format = fmt_map[col]


def auto_width(ws, min_width=10, max_width=28):
    for col_cells in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            if cell.value is not None:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = max(min_width, min(max_len + 3, max_width))


def freeze_and_filter(ws, freeze_cell="A2"):
    ws.freeze_panes = freeze_cell
    ws.auto_filter.ref = ws.dimensions


# ---------------------------------------------------------------------------
# Sheet Writers
# ---------------------------------------------------------------------------

def write_summary_sheet(wb, summary):
    ws = wb.create_sheet("Summary")
    ws.sheet_properties.tabColor = TAB_BLUE

    headers = ["Metric"] + list(summary.keys())
    for c, h in enumerate(headers, 1):
        ws.cell(row=1, column=c, value=h)
    style_header_row(ws, 1, len(headers))

    metric_formats = {
        "ALCC (M EUR/yr)": NUM_FMT_DEC2,
        "LCOE (EUR/MWh)": NUM_FMT_DEC2,
        "Emissions (tCO2/yr)": NUM_FMT_INT,
        "Avoided Emissions (tCO2/yr)": NUM_FMT_INT,
        "Peak Import (MW)": NUM_FMT_DEC1,
        "BESS Cycles/yr": NUM_FMT_INT,
        "Payback (years)": NUM_FMT_DEC1,
        "Annual Savings (EUR)": NUM_FMT_INT,
    }

    metrics = list(next(iter(summary.values())).keys())
    for r, metric in enumerate(metrics, 2):
        display_metric = (metric
                          .replace("€", "EUR")
                          .replace("₂", "2"))
        ws.cell(row=r, column=1, value=display_metric)
        ws.cell(row=r, column=1).border = BORDER
        ws.cell(row=r, column=1).font = Font(bold=True)
        ws.cell(row=r, column=1).alignment = Alignment(horizontal="left")

        fmt = metric_formats.get(display_metric, NUM_FMT_DEC2)
        for c, scenario in enumerate(summary.keys(), 2):
            val = summary[scenario][metric]
            if val == "N/A":
                ws.cell(row=r, column=c, value="N/A")
            elif metric == "Payback (years)" and isinstance(val, float) and val > 50:
                ws.cell(row=r, column=c, value=">50")
            else:
                ws.cell(row=r, column=c, value=val)
            style_data_cell(ws, r, c, fmt if val != "N/A" and not isinstance(val, str) else None)

    ws.freeze_panes = "B2"
    auto_width(ws)


def write_assumptions_sheet(wb):
    ws = wb.create_sheet("Assumptions")
    ws.sheet_properties.tabColor = TAB_BLUE

    headers = ["Parameter", "Thesis Value", "Updated Value", "Unit", "Source"]
    for c, h in enumerate(headers, 1):
        ws.cell(row=1, column=c, value=h)
    style_header_row(ws, 1, len(headers))

    rows = [
        ("DC Load", 10, 10, "MW", "Group agreement"),
        ("Grid Connection (MIC)", 20, 20, "MW", "2x DC load for BESS charging headroom"),
        ("Grid CO2 Intensity", 225, 225, "gCO2/kWh", "SEAI"),
        ("TOU Tariff - Day", 65.03, 65.03, "EUR/MWh", "Thesis Table 3.2"),
        ("TOU Tariff - Peak", 66.40, 66.40, "EUR/MWh", "Thesis Table 3.2"),
        ("TOU Tariff - Night", 53.53, 53.53, "EUR/MWh", "Thesis Table 3.2"),
        ("Grid Subscription Fee", 4_403_376, 4_403_376, "EUR/yr", "Thesis Table 3.3"),
        ("BESS Power CapEx", 675, 221.4, "EUR/kW", "NREL ATB 2025 / NREL ATB 2024 Adv."),
        ("BESS Energy CapEx", 165.6, 239.6, "EUR/kWh", "NREL ATB 2025 / NREL ATB 2024 Adv."),
        ("BESS Power O&M", 16.9, 5.5, "EUR/kW-yr", "NREL ATB 2025 / NREL ATB 2024 Adv."),
        ("BESS Energy O&M", 4.11, 6.0, "EUR/kWh-yr", "NREL ATB 2025 / NREL ATB 2024 Adv."),
        ("Round-Trip Efficiency", "95%", "95%", "", "Thesis Table 3.7"),
        ("BESS Lifetime", 15, 15, "years", "Thesis Table 3.7"),
        ("WACC", "8%", "8%", "", "Thesis assumption"),
        ("Capital Recovery Factor", round(AF_BESS, 5), round(AF_BESS, 5), "", "Computed"),
        ("Initial SoC", "50%", "50%", "of E_BESS", "Thesis assumption"),
        ("Spot Price Source", "Synthetic", "Synthetic", "", "Calibrated to 2023 I-SEM stats"),
    ]
    for r, row_data in enumerate(rows, 2):
        for c, val in enumerate(row_data, 1):
            ws.cell(row=r, column=c, value=val)
            style_data_cell(ws, r, c)
        ws.cell(row=r, column=1).alignment = Alignment(horizontal="left")

    ws.freeze_panes = "A2"
    auto_width(ws)


def write_baseline_sheet(wb, spot, tou):
    ws = wb.create_sheet("Baseline")
    ws.sheet_properties.tabColor = TAB_BLUE

    headers = ["Month", "Hours", "Avg Spot (EUR/MWh)", "Avg TOU (EUR/MWh)",
               "Import (MWh)", "Energy Cost (EUR)", "Emissions (tCO2)"]
    for c, h in enumerate(headers, 1):
        ws.cell(row=1, column=c, value=h)
    style_header_row(ws, 1, len(headers))

    months_arr = build_calendar()[0]
    total_hours = 0
    total_import = 0.0
    total_cost = 0.0
    total_emissions = 0.0

    for m in range(12):
        mask = months_arr == (m + 1)
        n_hours = int(mask.sum())
        avg_spot = float(np.mean(spot[mask]))
        avg_tou = float(np.mean(tou[mask]))
        import_mwh = D_MW * n_hours
        energy_cost = D_MW * float(np.sum((spot + tou)[mask]))
        emissions = import_mwh * CARBON_INTENSITY / 1000

        total_hours += n_hours
        total_import += import_mwh
        total_cost += energy_cost
        total_emissions += emissions

        row = m + 2
        vals = [MONTH_NAMES[m], n_hours, avg_spot, avg_tou, import_mwh, energy_cost, emissions]
        fmts = [None, NUM_FMT_INT, NUM_FMT_DEC2, NUM_FMT_DEC2, NUM_FMT_INT, NUM_FMT_INT, NUM_FMT_INT]
        for c, (val, fmt) in enumerate(zip(vals, fmts), 1):
            ws.cell(row=row, column=c, value=val)
            style_data_cell(ws, row, c, fmt)
        ws.cell(row=row, column=1).alignment = Alignment(horizontal="left")

    total_row = 14
    total_vals = ["TOTAL", total_hours, "", "", total_import, total_cost, total_emissions]
    total_fmts = {2: NUM_FMT_INT, 5: NUM_FMT_INT, 6: NUM_FMT_INT, 7: NUM_FMT_INT}
    for c, val in enumerate(total_vals, 1):
        ws.cell(row=total_row, column=c, value=val if val != "" else None)
    style_total_row(ws, total_row, len(headers), total_fmts)

    freeze_and_filter(ws)
    auto_width(ws)


def write_bess_sizing_sheet(wb, bess_results_thesis, bess_results_updated, baseline_cost):
    ws = wb.create_sheet("BESS Sizing")
    ws.sheet_properties.tabColor = TAB_GREEN

    headers = ["Config", "P (MW)", "Tau (h)", "E (MWh)",
               "Annual BESS Cost (EUR) [Thesis]", "Total Cost (EUR) [Thesis]", "Savings (EUR) [Thesis]",
               "Annual BESS Cost (EUR) [Updated]", "Total Cost (EUR) [Updated]", "Savings (EUR) [Updated]",
               "Payback (yr) [Updated]"]
    for c, h in enumerate(headers, 1):
        ws.cell(row=1, column=c, value=h)
    style_header_row(ws, 1, len(headers))

    best_updated_key = min(bess_results_updated, key=lambda k: bess_results_updated[k]["total_cost"])

    row = 2
    for P in P_OPTIONS:
        for tau in TAU_OPTIONS:
            key = (P, tau)
            if key not in bess_results_thesis:
                continue
            rt = bess_results_thesis[key]
            ru = bess_results_updated[key]
            E = P * tau
            savings_t = baseline_cost - rt["total_cost"]
            savings_u = baseline_cost - ru["total_cost"]
            capex_u = BESS_UPDATED["c_P_cap"] * P * 1000 + BESS_UPDATED["c_E_cap"] * E * 1000
            payback_u = capex_u / savings_u if savings_u > 0 else None

            payback_display = payback_u if payback_u is not None else "N/A"

            vals = [f"{P}MW/{E}MWh", P, tau, E,
                    rt["bess_annual_cost"], rt["total_cost"], savings_t,
                    ru["bess_annual_cost"], ru["total_cost"], savings_u,
                    payback_display]
            fmts = [None, NUM_FMT_INT, NUM_FMT_INT, NUM_FMT_INT,
                    NUM_FMT_INT, NUM_FMT_INT, NUM_FMT_INT,
                    NUM_FMT_INT, NUM_FMT_INT, NUM_FMT_INT,
                    NUM_FMT_DEC1]
            for c, (val, fmt) in enumerate(zip(vals, fmts), 1):
                ws.cell(row=row, column=c, value=val)
                style_data_cell(ws, row, c, fmt if not isinstance(val, str) else None)

            if key == best_updated_key:
                for c in range(1, len(headers) + 1):
                    ws.cell(row=row, column=c).fill = HIGHLIGHT_FILL
            row += 1

    freeze_and_filter(ws)
    auto_width(ws)


def write_dispatch_sheet(wb, spot, tou, dispatch):
    ws = wb.create_sheet("Optimal Dispatch")
    ws.sheet_properties.tabColor = TAB_GREEN

    headers = ["Hour", "Date", "Spot (EUR/MWh)", "TOU (EUR/MWh)",
               "Grid Import (MW)", "Charge (MW)", "Discharge (MW)", "SoC (MWh)", "Hourly Cost (EUR)"]
    for c, h in enumerate(headers, 1):
        ws.cell(row=1, column=c, value=h)
    style_header_row(ws, 1, len(headers))

    fmts = [NUM_FMT_INT, None, NUM_FMT_DEC2, NUM_FMT_DEC2,
            NUM_FMT_DEC2, NUM_FMT_DEC2, NUM_FMT_DEC2, NUM_FMT_DEC2, NUM_FMT_DEC2]

    for t in range(T):
        row = t + 2
        dt = hour_index_to_datetime(t)
        cost = dispatch["grid_import"][t] * (spot[t] + tou[t])
        vals = [t, dt.strftime("%Y-%m-%d %H:%M"), spot[t], tou[t],
                dispatch["grid_import"][t], dispatch["charge"][t],
                dispatch["discharge"][t], dispatch["soc"][t], cost]
        for c, (val, fmt) in enumerate(zip(vals, fmts), 1):
            cell = ws.cell(row=row, column=c, value=val)
            if fmt:
                cell.number_format = fmt

    ws.freeze_panes = "A2"
    for col_idx, width in enumerate([7, 18, 14, 14, 14, 12, 12, 12, 14], 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width


def write_dr_sheet(wb, dr_result):
    ws = wb.create_sheet("DR Analysis")
    ws.sheet_properties.tabColor = TAB_ORANGE

    headers = ["Event #", "Day of Year", "Date", "Start Hour",
               "Duration (h)", "Baseline Import (MW)", "Actual Import (MW)",
               "Curtailed (MWh)", "Capacity Claim (MW)"]
    for c, h in enumerate(headers, 1):
        ws.cell(row=1, column=c, value=h)
    style_header_row(ws, 1, len(headers))

    events = dr_result["events"]
    fmts = [NUM_FMT_INT, NUM_FMT_INT, None, None, NUM_FMT_INT,
            NUM_FMT_DEC2, NUM_FMT_DEC2, NUM_FMT_DEC2, NUM_FMT_DEC2]

    for idx, ev in enumerate(events):
        row = idx + 2
        dt = hour_index_to_datetime(ev["start_hour"])
        actual_import = D_MW - ev["curtailed_MWh"] / ev["duration_h"] if ev["duration_h"] > 0 else D_MW

        vals = [idx + 1, ev["day"], dt.strftime("%Y-%m-%d"),
                dt.strftime("%H:%M"), ev["duration_h"],
                ev["baseline_import_MW"], actual_import,
                ev["curtailed_MWh"], ev["capacity_claim_MW"]]
        for c, (val, fmt) in enumerate(zip(vals, fmts), 1):
            ws.cell(row=row, column=c, value=val)
            style_data_cell(ws, row, c, fmt)

    total_row = len(events) + 2
    ws.cell(row=total_row, column=1, value="TOTAL")
    ws.cell(row=total_row, column=8, value=sum(e["curtailed_MWh"] for e in events))
    ws.cell(row=total_row, column=9, value=float(np.mean([e["capacity_claim_MW"] for e in events])))
    total_fmts = {8: NUM_FMT_DEC2, 9: NUM_FMT_DEC2}
    style_total_row(ws, total_row, len(headers), total_fmts)

    # DSU Revenue Summary — separate section with its own header
    summary_start = total_row + 2
    ws.cell(row=summary_start, column=1, value="DSU Revenue Summary")
    ws.cell(row=summary_start, column=1).font = Font(bold=True, size=12)

    sub_headers = ["Metric", "Value"]
    for c, h in enumerate(sub_headers, 1):
        ws.cell(row=summary_start + 1, column=c, value=h)
    style_header_row(ws, summary_start + 1, 2)

    enrolled = dr_result.get("enrolled_capacity_MW", 0)
    eligible = dr_result.get("dsu_eligible", False)
    labels = [
        ("Enrolled Capacity (MW)", enrolled, NUM_FMT_DEC2),
        ("DSU Eligible (>=4 MW)", "Yes" if eligible else "No", None),
        ("EirGrid Capacity Payment (EUR)", dr_result.get("annual_cap_payment", 0), NUM_FMT_INT),
        ("Total DR Revenue (EUR)", dr_result["dr_revenue"], NUM_FMT_INT),
        ("Annual Peak Discharge (MWh)", dr_result.get("annual_peak_discharge_MWh", 0), NUM_FMT_INT),
        ("BESS Deficit vs Baseline (EUR)", dr_result["bess_deficit"], NUM_FMT_INT),
        ("Net DR Benefit (EUR)", dr_result["dr_net"], NUM_FMT_INT),
    ]
    for i, (label, val, fmt) in enumerate(labels):
        r = summary_start + 2 + i
        ws.cell(row=r, column=1, value=label)
        ws.cell(row=r, column=1).border = BORDER
        ws.cell(row=r, column=1).alignment = Alignment(horizontal="left")
        ws.cell(row=r, column=2, value=val)
        style_data_cell(ws, r, 2, fmt if not isinstance(val, str) else None)

    freeze_and_filter(ws, "A2")
    auto_width(ws)


def write_cost_comparison_sheet(wb, spot, tou, dispatch):
    ws = wb.create_sheet("Cost Comparison")
    ws.sheet_properties.tabColor = TAB_GREEN

    headers = ["Month", "Grid-Only Cost (EUR)", "BESS Import Cost (EUR)",
               "Savings (EUR)", "Savings (%)",
               "Grid-Only Emissions (tCO2)", "BESS Emissions (tCO2)", "Emissions Delta (tCO2)"]
    for c, h in enumerate(headers, 1):
        ws.cell(row=1, column=c, value=h)
    style_header_row(ws, 1, len(headers))

    months_arr = build_calendar()[0]
    totals = [0.0] * 7

    for m in range(12):
        mask = months_arr == (m + 1)
        total_price = spot + tou
        baseline_cost = D_MW * float(np.sum(total_price[mask]))
        bess_cost = float(np.sum(dispatch["grid_import"][mask] * total_price[mask]))
        savings = baseline_cost - bess_cost
        pct = savings / baseline_cost if baseline_cost > 0 else 0
        base_em = D_MW * int(mask.sum()) * CARBON_INTENSITY / 1000
        bess_em = float(np.sum(dispatch["grid_import"][mask])) * CARBON_INTENSITY / 1000

        totals[0] += baseline_cost
        totals[1] += bess_cost
        totals[2] += savings
        totals[4] += base_em
        totals[5] += bess_em
        totals[6] += base_em - bess_em

        row = m + 2
        vals = [MONTH_NAMES[m], baseline_cost, bess_cost, savings, pct,
                base_em, bess_em, base_em - bess_em]
        fmts = [None, NUM_FMT_INT, NUM_FMT_INT, NUM_FMT_INT, NUM_FMT_PCT,
                NUM_FMT_DEC1, NUM_FMT_DEC1, NUM_FMT_DEC1]
        for c, (val, fmt) in enumerate(zip(vals, fmts), 1):
            ws.cell(row=row, column=c, value=val)
            style_data_cell(ws, row, c, fmt)
        ws.cell(row=row, column=1).alignment = Alignment(horizontal="left")

    total_row = 14
    totals[3] = totals[2] / totals[0] if totals[0] > 0 else 0
    total_vals = ["TOTAL", totals[0], totals[1], totals[2], totals[3],
                  totals[4], totals[5], totals[6]]
    total_fmts_map = {2: NUM_FMT_INT, 3: NUM_FMT_INT, 4: NUM_FMT_INT, 5: NUM_FMT_PCT,
                      6: NUM_FMT_DEC1, 7: NUM_FMT_DEC1, 8: NUM_FMT_DEC1}
    for c, val in enumerate(total_vals, 1):
        ws.cell(row=total_row, column=c, value=val)
    style_total_row(ws, total_row, len(headers), total_fmts_map)

    freeze_and_filter(ws)
    auto_width(ws)


def write_beta_sweep_sheet(wb, beta_sweep, baseline_cost):
    ws = wb.create_sheet("Beta Sensitivity")
    ws.sheet_properties.tabColor = TAB_PURPLE

    # Table 1: Optimal config per beta
    ws.cell(row=1, column=1, value="Optimal Configuration per Discount Factor")
    ws.cell(row=1, column=1).font = Font(bold=True, size=12)

    headers = ["Beta", "Discount Formula", "Optimal Config",
               "Adj. DR Revenue (EUR)", "Adj. Net Cost (EUR)", "Savings vs Grid-Only (EUR)"]
    header_row = 2
    for c, h in enumerate(headers, 1):
        ws.cell(row=header_row, column=c, value=h)
    style_header_row(ws, header_row, len(headers))

    for r, entry in enumerate(beta_sweep["sweep_results"], header_row + 1):
        beta = entry["beta"]
        bk = entry["best_key"]
        cfg = entry["configs"][bk]
        vals = [
            beta,
            f"1 - {beta:.2f} * (P / 10)",
            f"{cfg['P_MW']}MW/{cfg['E_MWh']}MWh",
            cfg["adjusted_dr_revenue"],
            cfg["adjusted_net_cost"],
            baseline_cost - cfg["adjusted_net_cost"],
        ]
        fmts = [NUM_FMT_DEC2, None, None, NUM_FMT_INT, NUM_FMT_INT, NUM_FMT_INT]
        for c, (val, fmt) in enumerate(zip(vals, fmts), 1):
            ws.cell(row=r, column=c, value=val)
            style_data_cell(ws, r, c, fmt)

    # Table 2: Full matrix
    gap_row = header_row + len(beta_sweep["sweep_results"]) + 3
    ws.cell(row=gap_row, column=1, value="Adjusted Net Cost by Configuration and Beta (EUR)")
    ws.cell(row=gap_row, column=1).font = Font(bold=True, size=12)

    matrix_header_row = gap_row + 1
    beta_values = beta_sweep["beta_values"]
    ws.cell(row=matrix_header_row, column=1, value="Config")
    for j, beta in enumerate(beta_values):
        ws.cell(row=matrix_header_row, column=j + 2, value=f"B={float(beta):.2f}")
    style_header_row(ws, matrix_header_row, len(beta_values) + 1)

    data_row = matrix_header_row + 1
    for P in P_OPTIONS:
        for tau in TAU_OPTIONS:
            key = (P, tau)
            E = P * tau
            ws.cell(row=data_row, column=1, value=f"{P}MW/{E}MWh")
            style_data_cell(ws, data_row, 1)
            for j, entry in enumerate(beta_sweep["sweep_results"]):
                if key in entry["configs"]:
                    val = entry["configs"][key]["adjusted_net_cost"]
                    ws.cell(row=data_row, column=j + 2, value=val)
                    style_data_cell(ws, data_row, j + 2, NUM_FMT_INT)
            data_row += 1

    ws.freeze_panes = "B3"
    auto_width(ws, min_width=12)


# ---------------------------------------------------------------------------
# Workbook Assembly
# ---------------------------------------------------------------------------

def generate_workbook(summary, spot, tou, bess_results_thesis, bess_results_updated,
                      baseline, optimal_dispatch, dr_result, baseline_cost, beta_sweep=None):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    write_summary_sheet(wb, summary)
    write_assumptions_sheet(wb)
    write_baseline_sheet(wb, spot, tou)
    write_bess_sizing_sheet(wb, bess_results_thesis, bess_results_updated, baseline_cost)
    write_dispatch_sheet(wb, spot, tou, optimal_dispatch)
    write_dr_sheet(wb, dr_result)
    write_cost_comparison_sheet(wb, spot, tou, optimal_dispatch)

    if beta_sweep is not None:
        write_beta_sweep_sheet(wb, beta_sweep, baseline_cost)

    wb.save(OUTPUT_XLSX)
    print(f"  Workbook saved: {OUTPUT_XLSX}")


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    print("Loading results from", RESULTS_PKL)
    with open(RESULTS_PKL, "rb") as f:
        results = pickle.load(f)

    spot = results["spot"]
    tou = results["tou"]
    baseline = results["baseline"]
    optimal_dispatch = results["optimal_dispatch"]
    dr_result = results["dr_result"]
    summary = results["summary"]
    bess_results_thesis = results["bess_results_thesis"]
    bess_results_updated = results["bess_results_updated"]

    beta_sweep = results.get("beta_sweep")

    print("Generating Excel workbook...")
    generate_workbook(summary, spot, tou, bess_results_thesis, bess_results_updated,
                      baseline, optimal_dispatch, dr_result, baseline["total_cost"],
                      beta_sweep=beta_sweep)

    print(f"Done. Workbook saved to {OUTPUT_XLSX}")


if __name__ == "__main__":
    main()
